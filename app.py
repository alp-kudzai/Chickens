#import xlwt
import openpyxl
import pandas
import re
import sys
import datetime
#Overview
#Creating an app that tracks chicken sales and cost as spreadsheet data.


PRICE = 90
# TODO
# I will probably need something that check whether this customer owes us and carries the debt over to the status automatically
# find out what error is thrown when you try to save with current workbook open and catch it, ANS ==> PermissionError

#A function that saves a "Workbook" aka spreadsheet to current directory
def save_sheet(name:str, workbook:callable):
    if (".xlsx" not in name):
        name += ".xlsx"
    workbook.save(filename=name)
#A function that loads a spreadsheet given a name
def load_sheet(name:str):
    '''
    Given a name of a file we load a the spreadsheet
    '''
    if (".xlsx" not in name):
        name += ".xlsx"
    workbook = openpyxl.load_workbook(filename=name)
    return workbook

#a func that creates the "sheet1 = sales" default rows
def create_workbook():
    workbook = openpyxl.Workbook()
    sales_sheet, cost_sheet = workbook.active, workbook.create_sheet("Costs", 1)
    #create the initial column heading
    sales_sheet.title = "Sales"
    sales_sheet.append([
        "Date",
        "Name",
        "No. of Sales",
        "Nature of Sale",
        "Rand Amount",
        "Status"
    ])
    cost_sheet.append([
        "Items",
        "Number",
        "Unit",
        "Cost in Rand"
    ])
    # It is creating a default sheet with no name, I dont what that
    #del workbook["sheet"]
    #cant delete the default
    return workbook

def create_dfs():
    sales_heading = [
        "Date",
        "Name",
        "No. of Sales",
        "Nature of Sale",
        "Rand Amount",
        "Status"
    ]
    costs_headings = [
        "Items",
        "Number",
        "Unit",
        "Cost in Rand"
    ]
    sales_df, costs_df = pandas.DataFrame(columns=sales_heading), pandas.DataFrame(columns=costs_headings)
    return sales_df, costs_df
# a functions that takes a sheet name and returns the active sheet
def activate_sheet(name:str, workbook:callable):
    return workbook[name]

def assign_status(nature:str, number:int):
    assert nature == "credit" or nature == "cash"
    if nature == "credit": return -(number * PRICE)
    else: return 0

# function that appends new data to the active sheet
def append_sales(name:str, number:str, nature:str, sales_sheet:callable, date=datetime.date.today()):
    if sales_sheet.title == "Sales":
        #convert number to interger
        number = int(number)
        status = assign_status(nature, number)
        regx_pattern = r"^\d{4}-(0[1-9]|1[0-2])-(0[1-9]|[12][0-9]|3[01])$"
        #converting datetime object to string
        date = str(date)
        #making sure that the date is in the right format YYY-MM-DD with regx
        final_date = re.search(regx_pattern, date)
        if final_date == None:
            raise ValueError("The date is in the wrong format, make sure it is YYYY-MM-DD")
        date_string = final_date.string
        sales_sheet.append([
            date_string,
            name,
            number,
            nature,
            number*PRICE,
            status,
        ])
    else:
        raise ValueError("This is not the Sales Sheet")
    return sales_sheet

def append_sales_df(name:str, number:str, nature:str,sales_df:callable, date=datetime.date.today()):
    number = int(number)
    status = assign_status(nature, number)
    regx_pattern = r"^\d{4}-(0[1-9]|1[0-2])-(0[1-9]|[12][0-9]|3[01])$"
    #converting datetime object to string
    date = str(date)
    #making sure that the date is in the right format YYY-MM-DD with regx
    final_date = re.search(regx_pattern, date)
    if final_date == None:
        raise ValueError("The date is in the wrong format, make sure it is YYYY-MM-DD")
    data = [date, name, number, nature, number*PRICE, status]
    sales_df.loc[len(sales_df)] = data # append the data via an index that is dependant on the length of the dataframe. So 0 len we place it at 0 index
    return sales_df

    
#A function that returns a dataframe of a sheet given the name and a workbook instance
def ws2df(name:str, workbook:callable):
    '''
    A function that returns a dataframe of a sheet given the name and a workbook instance.

    '''
    ws = activate_sheet(name, workbook)
    data = ws.values
    heading = next(data)[0:]
    return pandas.DataFrame(data, columns=heading)

#an alternative func that loads the excel sheet straight into a dataframe
def load_dfs(fname:str):
    if (".xlsx" not in fname):
        fname += ".xlsx"
    return pandas.read_excel(fname, sheet_name="Sales"), pandas.read_excel(fname, sheet_name="Costs")
#Save dataframes to excel
def save_dfs(fname:str, sales_df:callable, costs_df:callable):
    try:
        if ".xlsx" not in fname:
            fname += ".xlsx"
        with pandas.ExcelWriter(fname) as writer:
            sales_df.to_excel(writer, "Sales", index=False)
            costs_df.to_excel(writer, "Costs", index=False)
    except PermissionError:
        print("[!] Permission Error: Make Sure that the file is close in another window!!")
        
#A search func, that goes over the names in the spreadsheet an
def name_search_df(name:str, df:callable):
    '''
    1ST ATTEMPT
    We search the given dataframe for the name and return a dataframe
    '''
    return df.loc[df["Name"] == name]


# a functions that auto carries debt over from the last occurance of the given name. This will be searching in the data frame but 


def main():
    print('Chicken App, Welcome to Beta!\n\n')
    print('Current Options: load <file-name>; saveas <file-name>; save; create; sale <date> <name> <no.> <nature>; options;exit\n\n')
    run = True
    while run:
        user_input = input(">>> ")
        assert type(user_input) == str
        if "exit" in user_input:
            # careful to save before exit
            run = False
            sys.exit()
        elif "load" in user_input:
            fname = user_input.split(' ')[1] #get the filename from user input e.g load asd.xlsx
            try:
                open_workbook = load_sheet(fname)
                print('[*] Successful load\n')
            except:
                print("[!] Failed to load workbook!\n")
        elif "saveas" in user_input:
            fname = user_input.split(' ')[1]
            try:
                save_sheet(fname, open_workbook)
                print('[*] Successful save\n')
            except:
                print('[!] Failed to save workbook!\n')
        elif "save" in user_input:
            try:
                save_sheet(fname, open_workbook)
                print('[*] Successful save\n')
            except:
                print('[!] Failed to save workbook!\n')
        elif "create" in user_input:
            open_workbook = create_workbook()
            print('[*] Successfully created a new spreadsheet')
            print('[!] ITS HOT. Make sure to save!')
        elif "sale" in user_input:
            # data required order:
            # date, name, no. of chickens, nature, amount is global therefore chickens*PRICE, status
            # date is either given & if not given then we assume it from current date
            # we discard the first arg = sale 
            data = user_input.split(' ')[1:]
            sales_sheet = activate_sheet("Sales", open_workbook)
            if data.__len__() == 4:
                sales_sheet = append_sales(data[1], data[2], data[3], sales_sheet, date=data[0])
            elif data.__len__() == 3:
                sales_sheet = append_sales(data[0], data[1], data[2], sales_sheet)
            else:
                # Try fail gracefully
                raise ValueError("The user input doesnt contain the required amount of items! View options")
            print("[*] Successfully appended data to Sales Sheet.")
            print("[!] ITS HOT. Save before closing!")
        else:
            print('[!] Input Error, view Options: load <file-name>; save <file-name>; create; exit\n\n ')
            
if __name__ == '__main__':
    main()
                
